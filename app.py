# openpyxl to read xlsx files
import openpyxl
import csv
import os
from openpyxl import load_workbook

def workbook_list():
    #loop through xlsx files and add to array
    directory = r'C:\PythonParser\mondayAccounts'
    client_sheets = []
    for filename in os.listdir(directory):
        if filename.endswith('.xlsx'):
            client_sheets.append(filename)
        else:
            continue
    return client_sheets

def split_fields(wb):
        #split by line
        worksheet = wb.worksheets[0]
        client_name = [worksheet['A1'].value]
        first_split = worksheet['A2'].value.split('\n')

        #split by field, check for colon    
        for b in range (len(first_split)):
            if ':' in first_split[b]:
                first_split[b] = first_split[b].split(':', 1)
           
        #create titles and fields
        field_titles = []
        field_values = []
    
        for b in range (len(first_split)): 
            #if list item is a string and not another list
            if isinstance(first_split[b], str):
                first_split[b] = [first_split[b], first_split[b]]                   
            field_titles.append(first_split[b][0])
            field_values.append(first_split[b][1])
        
        return field_titles, field_values, client_name
       
def write_values(titles, values = [], is_first = False):
    if is_first:
        write = 'w'
    else:
        write = 'a'
    with open('C:\\PythonParser\\newclientwb.csv', write, encoding = 'UTF8') as f:
        writer = csv.writer(f)
        writer.writerow(titles)
        if values == []:
            pass
        else:
            writer.writerow(values)

def services_info(wb):
    worksheet = wb.worksheets[0]
    amount_of_rows = worksheet.max_row
    for i in range(amount_of_rows):
        row = i + 1
        if worksheet.cell(row, 1).value == 'SERVICES PROVIDED':
            services_name = []
            services_status = []
            services_date = []
            services_notes = []
            row = i + 3
            items = 0
            while worksheet.cell(row, 1).value:
                services_name.append(worksheet.cell(row, 1).value)
                services_status.append(worksheet.cell(row, 2).value)
                services_date.append(worksheet.cell(row, 3).value)
                services_notes.append(worksheet.cell(row, 4).value)
                row += 1
                items += 1
            if items > 0:
                return services_name, services_status, services_date, services_notes
            else:
                return [], [], [], []
    return [], [], [], []
        
        
client_sheets = workbook_list()


is_first = True
for i in client_sheets:
    workbook = load_workbook(filename = 'C:\PythonParser\mondayAccounts\\' + i)
    info_titles, info_values, client_name = split_fields(workbook)
    name, status, date, notes = services_info(workbook)
    
    #add columns to each list for name and description
    info_titles = [client_name[0], 'Info Titles'] + info_titles
    info_values = [client_name[0], 'Info Values'] + info_values
    name = [client_name[0], 'Service Name'] + name
    status = [client_name[0], 'Service Status'] + status
    date = [client_name[0], 'Date Modified'] + date
    notes = [client_name[0], 'Notes, if any'] + notes

    write_values(client_name)
    write_values(info_titles, info_values)
    write_values(name, status)
    write_values(date, notes)
    if is_first:
        is_first = False      
